from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from typing import Any
import os


class ExcelMapper:
    def __init__(self, filename: str, sheetIdx: int=0, newOpenAble: bool = False):
        """ファイルの読み込み・作成を行います．"""
        self.wb: Workbook = None
        if os.path.exists(filename):
            self.wb: Workbook = load_workbook(filename=filename)
        elif newOpenAble:
            self.wb: Workbook = Workbook()
        else:
            raise Exception()
        self.ws: Worksheet = self.openSheet(sheetIdx=sheetIdx)

    def openSheet(self, sheetIdx: int)->Worksheet:
        """シートを開きます"""
        return self.wb.worksheets[sheetIdx]

    def getValue(self, row: int, col: int)->Any:
        """値を取得します"""
        return self.ws.cell(row+1, col+1).value

    def setValue(self, row: int, col: int, value: Any):
        """値を設定します"""
        cell: Cell = self.ws.cell(row+1, col+1)
        cell.value = value

    def saveData(self, filename: str):
        """ファイルを保存します．"""
        self.wb.save(filename)

